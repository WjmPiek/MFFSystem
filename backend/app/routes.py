import csv
import io
import os
import secrets
from datetime import datetime, timedelta

from flask import Blueprint, current_app, g, jsonify, request
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from .extensions import db
from .models import MemberRecord, PaymentTransaction, User
from .permissions import auth_required, create_auth_token, normalize_role, roles_required
from .services.statement_parser import StatementParseError, StatementParser

api = Blueprint('api', __name__)
RESET_TOKENS = {}

MEMBER_CATEGORY_MAP = {
    'ins members': 'ins_members',
    'ins_members': 'ins_members',
    'ins-members': 'ins_members',
    'membership club': 'membership_club',
    'membership_club': 'membership_club',
    'membership-club': 'membership_club',
    'society': 'society',
}
DISPLAY_CATEGORY = {
    'ins_members': 'Ins Members',
    'membership_club': 'Membership Club',
    'society': 'Society',
}
MEMBER_FIELDS = [
    'member_number', 'first_name', 'last_name', 'full_name', 'email', 'phone', 'join_date', 'status', 'organisation_name', 'notes'
]


def _json():
    return request.get_json(silent=True) or {}


def _seed_sample_payments():
    if PaymentTransaction.query.count() > 0:
        return
    samples = [
        PaymentTransaction(payer_name='Alpha Foods', reference='ALPHA-1001', amount=2400.00, franchise_name='Pretoria East', statement_filename='march-statement.pdf', status='unmatched'),
        PaymentTransaction(payer_name='Beta Traders', reference='BETA-1002', amount=1675.50, franchise_name='Pretoria North', statement_filename='march-statement.pdf', status='allocated', allocated_to='Franchise Fee'),
        PaymentTransaction(payer_name='Gamma Supplies', reference='GAMMA-1003', amount=980.25, franchise_name='Centurion', statement_filename='march-statement.pdf', status='edited', notes='Reference corrected by admin'),
    ]
    db.session.add_all(samples)
    db.session.commit()


def _normalize_member_category(value):
    key = (value or '').strip().lower()
    return MEMBER_CATEGORY_MAP.get(key, key if key in DISPLAY_CATEGORY else None)


def _clean_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _normalize_header(value):
    return _clean_text(value).lower().replace('&', 'and').replace('-', ' ').replace('_', ' ')


def _build_member_payload(row):
    member_number = row.get('member_number') or row.get('member_id') or row.get('membership_number') or row.get('member no') or row.get('member no.') or row.get('number') or ''
    first_name = row.get('first_name') or row.get('name') or ''
    last_name = row.get('last_name') or row.get('surname') or ''
    full_name = row.get('full_name') or ' '.join(part for part in [first_name, last_name] if part).strip()
    email = row.get('email') or ''
    phone = row.get('phone') or row.get('cell') or row.get('mobile') or ''
    join_date = row.get('join_date') or row.get('join date') or row.get('date_joined') or ''
    status = row.get('status') or 'Active'
    organisation_name = row.get('organisation_name') or row.get('club_name') or row.get('society_name') or row.get('group_name') or ''
    notes = row.get('notes') or ''
    if not full_name:
        full_name = email or member_number
    return {
        'member_number': member_number or None,
        'first_name': first_name or None,
        'last_name': last_name or None,
        'full_name': full_name or None,
        'email': email or None,
        'phone': phone or None,
        'join_date': join_date or None,
        'status': status or None,
        'organisation_name': organisation_name or None,
        'notes': notes or None,
    }


def _validate_member_payload(category, payload, row_number):
    errors = []
    if not payload.get('full_name'):
        errors.append('Full name or first/surname is required.')
    if not payload.get('email') and not payload.get('member_number'):
        errors.append('Email or Member ID is required for duplicate detection.')
    if payload.get('email') and '@' not in payload['email']:
        errors.append('Email address is invalid.')
    return {'category': category, 'row_number': row_number, 'errors': errors, 'payload': payload}


def _rows_from_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(v) for v in rows[0]]
    out = []
    for values in rows[1:]:
        if not values or not any(v not in (None, '') for v in values):
            continue
        mapped = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            mapped[header] = _clean_text(values[idx]) if idx < len(values) else ''
        out.append(mapped)
    return out


def _parse_member_file(filename, file_bytes):
    lower = filename.lower()
    parsed = {}
    if lower.endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        for sheet_name in wb.sheetnames:
            category = _normalize_member_category(sheet_name)
            if not category:
                continue
            parsed[category] = _rows_from_sheet(wb[sheet_name])
    elif lower.endswith('.csv'):
        category = _normalize_member_category(request.form.get('category') or request.args.get('category') or '')
        if not category:
            raise ValueError('CSV imports require a category selection.')
        reader = csv.DictReader(io.StringIO(file_bytes.decode('utf-8-sig')))
        parsed[category] = [{_normalize_header(k): _clean_text(v) for k, v in row.items()} for row in reader]
    else:
        raise ValueError('Only .xlsx and .csv member imports are supported.')
    if not parsed:
        raise ValueError('No supported member sheets were found. Use Ins Members, Membership Club, and Society sheet names.')
    return parsed


def _member_import_preview(filename, file_bytes):
    parsed = _parse_member_file(filename, file_bytes)
    existing = {}
    for record in MemberRecord.query.all():
        existing.setdefault(record.category, set()).update(filter(None, [
            (record.email or '').strip().lower(),
            (record.member_number or '').strip().lower(),
        ]))
    preview_rows = []
    errors = []
    duplicates = []
    summary = {key: {'rows': 0, 'valid': 0, 'duplicates': 0, 'errors': 0} for key in DISPLAY_CATEGORY}
    for category, rows in parsed.items():
        for idx, raw in enumerate(rows, start=2):
            payload = _build_member_payload(raw)
            result = _validate_member_payload(category, payload, idx)
            summary[category]['rows'] += 1
            keys = [k for k in [
                (payload.get('email') or '').strip().lower(),
                (payload.get('member_number') or '').strip().lower(),
            ] if k]
            is_duplicate = any(k in existing.get(category, set()) for k in keys)
            if result['errors']:
                summary[category]['errors'] += 1
                errors.append({
                    'category': category,
                    'row_number': idx,
                    'full_name': payload.get('full_name'),
                    'errors': result['errors'],
                })
            elif is_duplicate:
                summary[category]['duplicates'] += 1
                summary[category]['valid'] += 1
                duplicates.append({
                    'category': category,
                    'row_number': idx,
                    'full_name': payload.get('full_name'),
                    'email': payload.get('email'),
                    'member_number': payload.get('member_number'),
                })
                preview_rows.append({'category': category, **payload, 'row_number': idx, 'action': 'update'})
            else:
                summary[category]['valid'] += 1
                preview_rows.append({'category': category, **payload, 'row_number': idx, 'action': 'create'})
    return {'rows': preview_rows, 'errors': errors, 'duplicates': duplicates, 'summary': summary}


def _upsert_member(category, payload, filename):
    query = None
    if payload.get('email'):
        query = MemberRecord.query.filter(db.func.lower(MemberRecord.email) == payload['email'].lower(), MemberRecord.category == category).first()
    if query is None and payload.get('member_number'):
        query = MemberRecord.query.filter(db.func.lower(MemberRecord.member_number) == payload['member_number'].lower(), MemberRecord.category == category).first()
    record = query or MemberRecord(category=category)
    for field in MEMBER_FIELDS:
        setattr(record, field, payload.get(field))
    record.source_filename = filename
    if query is None:
        db.session.add(record)
    return record, query is None


@api.get('/health')
def health_check():
    return jsonify({'status': 'ok'}), 200


@api.post('/auth/login')
def login():
    data = _json()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not email or not password:
        return jsonify({'error': 'Email and password are required.'}), 400

    user = User.query.filter(db.func.lower(User.email) == email).first()
    if not user or not user.is_active or not user.check_password(password):
        return jsonify({'error': 'Incorrect email or password.'}), 401

    user.role = normalize_role(user.role)
    token = create_auth_token(current_app.config['SECRET_KEY'], user)
    return jsonify({'token': token, 'user': user.to_dict()}), 200


@api.get('/auth/me')
@auth_required
def auth_me():
    return jsonify({'user': g.current_user.to_dict()}), 200


@api.post('/auth/request-reset')
def request_reset():
    data = _json()
    email = (data.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'Email is required.'}), 400

    user = User.query.filter(db.func.lower(User.email) == email).first()
    if user:
        token = secrets.token_urlsafe(24)
        RESET_TOKENS[token] = {
            'user_id': user.id,
            'expires_at': datetime.utcnow() + timedelta(hours=1),
        }
    else:
        token = None

    response = {'message': 'If the account exists, a password reset token has been created for testing.'}
    if current_app.config.get('ENV') != 'production' and token:
        response['reset_token'] = token
    return jsonify(response), 200


@api.post('/auth/reset-password')
def reset_password():
    data = _json()
    token = (data.get('token') or '').strip()
    new_password = data.get('new_password') or ''
    if not token or not new_password:
        return jsonify({'error': 'Token and new password are required.'}), 400

    payload = RESET_TOKENS.get(token)
    if not payload or payload['expires_at'] < datetime.utcnow():
        RESET_TOKENS.pop(token, None)
        return jsonify({'error': 'Reset token is invalid or expired.'}), 400

    user = User.query.get(payload['user_id'])
    if not user:
        RESET_TOKENS.pop(token, None)
        return jsonify({'error': 'User not found.'}), 404

    user.set_password(new_password)
    db.session.commit()
    RESET_TOKENS.pop(token, None)
    return jsonify({'message': 'Password reset successfully.'}), 200


@api.get('/users')
@roles_required('admin')
def get_users():
    users = User.query.order_by(User.id.desc()).all()
    return jsonify([user.to_dict() for user in users]), 200


@api.post('/users')
@roles_required('admin')
def create_user():
    data = _json()
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    role = normalize_role(data.get('role'))
    is_active = bool(data.get('is_active', True))

    if not name or not email or not password:
        return jsonify({'error': 'Name, email, and password are required.'}), 400

    existing_user = User.query.filter(db.func.lower(User.email) == email).first()
    if existing_user:
        return jsonify({'error': 'Email already exists.'}), 409

    user = User(name=name, email=email, role=role, is_active=is_active)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return jsonify(user.to_dict()), 201


@api.put('/users/<int:user_id>')
@roles_required('admin')
def update_user(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found.'}), 404

    data = _json()
    if 'name' in data:
        user.name = (data.get('name') or user.name).strip() or user.name
    if 'email' in data:
        new_email = (data.get('email') or '').strip().lower()
        if not new_email:
            return jsonify({'error': 'Email cannot be empty.'}), 400
        existing_user = User.query.filter(db.func.lower(User.email) == new_email, User.id != user_id).first()
        if existing_user:
            return jsonify({'error': 'Email already exists.'}), 409
        user.email = new_email
    if 'role' in data:
        user.role = normalize_role(data.get('role'))
    if 'is_active' in data:
        user.is_active = bool(data.get('is_active'))
    if data.get('password'):
        user.set_password(data['password'])

    db.session.commit()
    return jsonify(user.to_dict()), 200


@api.post('/users/<int:user_id>/reset-password')
@roles_required('admin')
def admin_reset_user_password(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found.'}), 404

    data = _json()
    new_password = data.get('new_password') or ''
    if not new_password:
        return jsonify({'error': 'New password is required.'}), 400

    user.set_password(new_password)
    db.session.commit()
    return jsonify({'message': 'Password updated successfully.'}), 200


@api.delete('/users/<int:user_id>')
@roles_required('admin')
def delete_user(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found.'}), 404
    if user.email.lower() == os.getenv('ADMIN_EMAIL', 'wjm@martinsdirect.com').strip().lower():
        return jsonify({'error': 'The seeded admin account cannot be deleted.'}), 400

    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted.'}), 200


@api.get('/members')
@roles_required('admin', 'franchisee')
def list_members():
    category = _normalize_member_category(request.args.get('category') or 'ins_members') or 'ins_members'
    records = MemberRecord.query.filter_by(category=category).order_by(MemberRecord.full_name.asc()).all()
    return jsonify({
        'category': category,
        'label': DISPLAY_CATEGORY.get(category, category),
        'members': [record.to_dict() for record in records],
    }), 200


@api.get('/members/import-template-info')
@roles_required('admin', 'franchisee')
def member_import_template_info():
    return jsonify({
        'template_path': '/templates/members-import-template.xlsx',
        'categories': [
            {'key': 'ins_members', 'label': 'Ins Members'},
            {'key': 'membership_club', 'label': 'Membership Club'},
            {'key': 'society', 'label': 'Society'},
        ],
        'required_columns': ['Member ID / Email', 'Name or First Name + Surname'],
        'notes': [
            'Use the provided workbook template and keep the sheet names unchanged.',
            'Preview checks duplicates, missing names, and invalid emails before saving.',
            'Commit imports create new records and update matching records by Email or Member ID within each category.',
        ],
    }), 200


@api.post('/members/import/preview')
@roles_required('admin', 'franchisee')
def preview_member_import():
    import_file = request.files.get('file') or next(iter(request.files.values()), None)
    if not import_file or not import_file.filename:
        return jsonify({'error': 'Choose an Excel or CSV file to preview.'}), 400
    try:
        preview = _member_import_preview(import_file.filename, import_file.read())
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify(preview), 200


@api.post('/members/import/commit')
@roles_required('admin', 'franchisee')
def commit_member_import():
    import_file = request.files.get('file') or next(iter(request.files.values()), None)
    if not import_file or not import_file.filename:
        return jsonify({'error': 'Choose an Excel or CSV file to import.'}), 400
    filename = secure_filename(import_file.filename) or 'members-import.xlsx'
    try:
        preview = _member_import_preview(filename, import_file.read())
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    created = 0
    updated = 0
    for row in preview['rows']:
        category = row['category']
        payload = {key: row.get(key) for key in MEMBER_FIELDS}
        _, is_created = _upsert_member(category, payload, filename)
        created += 1 if is_created else 0
        updated += 0 if is_created else 1
    db.session.commit()
    return jsonify({
        'message': f'Member import completed. {created} created, {updated} updated.',
        'created': created,
        'updated': updated,
        'duplicates_skipped': len(preview['duplicates']),
        'errors': preview['errors'],
        'summary': preview['summary'],
    }), 200


@api.get('/payments')
@roles_required('admin', 'franchisee')
def list_payments():
    _seed_sample_payments()
    payments = PaymentTransaction.query.order_by(PaymentTransaction.created_at.desc()).all()
    return jsonify([payment.to_dict() for payment in payments]), 200


@api.get('/payments/import-options')
@roles_required('admin', 'franchisee')
def payment_import_options():
    return jsonify({
        'supported_banks': ['nedbank', 'absa', 'fnb', 'standard_bank', 'capitec'],
        'supported_extensions': ['.pdf', '.csv', '.xlsx', '.xls', '.json'],
        'notes': [
            'PDF import works for text-based statements.',
            'CSV and Excel import support common export layouts from Nedbank, ABSA, FNB, Standard Bank, and Capitec.',
            'Bank selection is optional but improves parsing accuracy when the statement layout is ambiguous.',
        ],
    }), 200


@api.post('/payments/upload-statement')
@roles_required('admin', 'franchisee')
def upload_statement():
    parser = StatementParser()
    created = []

    if request.files:
        statement_file = request.files.get('statement') or next(iter(request.files.values()), None)
        if not statement_file or not statement_file.filename:
            return jsonify({'error': 'A statement file is required.'}), 400

        filename = secure_filename(statement_file.filename) or 'bank-statement'
        franchise_name = (request.form.get('franchise_name') or '').strip() or None
        bank_name = (request.form.get('bank_name') or '').strip() or None

        try:
            rows = parser.parse_file(filename, statement_file.read(), franchise_name=franchise_name, bank_name=bank_name)
        except StatementParseError as exc:
            return jsonify({'error': str(exc)}), 400

    else:
        data = _json()
        filename = (data.get('filename') or 'uploaded-statement.json').strip()
        rows = data.get('transactions') or []

    for row in rows:
        payer_name = (row.get('payer_name') or '').strip()
        reference = (row.get('reference') or '').strip()
        try:
            amount = float(row.get('amount') or 0)
        except (TypeError, ValueError):
            amount = 0
        if not payer_name or not reference or amount <= 0:
            continue
        transaction = PaymentTransaction(
            payer_name=payer_name,
            reference=reference,
            amount=amount,
            franchise_name=(row.get('franchise_name') or '').strip() or None,
            statement_filename=filename,
            status='unmatched',
            notes=(row.get('notes') or '').strip() or None,
        )
        db.session.add(transaction)
        created.append(transaction)

    if not created:
        return jsonify({'error': 'No valid payment transactions were found to import.'}), 400

    db.session.commit()
    return jsonify({
        'message': f'{len(created)} transaction(s) imported from statement.',
        'payments': [payment.to_dict() for payment in created],
    }), 201


@api.put('/payments/<int:payment_id>/allocate')
@roles_required('admin', 'franchisee')
def allocate_payment(payment_id):
    payment = PaymentTransaction.query.get(payment_id)
    if not payment:
        return jsonify({'error': 'Payment not found.'}), 404

    data = _json()
    payment.allocated_to = (data.get('allocated_to') or '').strip() or payment.allocated_to
    payment.status = 'allocated'
    payment.notes = (data.get('notes') or payment.notes)
    db.session.commit()
    return jsonify(payment.to_dict()), 200


@api.put('/payments/<int:payment_id>')
@roles_required('admin', 'franchisee')
def edit_payment(payment_id):
    payment = PaymentTransaction.query.get(payment_id)
    if not payment:
        return jsonify({'error': 'Payment not found.'}), 404

    data = _json()
    if 'payer_name' in data:
        payment.payer_name = (data.get('payer_name') or payment.payer_name).strip() or payment.payer_name
    if 'reference' in data:
        payment.reference = (data.get('reference') or payment.reference).strip() or payment.reference
    if 'amount' in data:
        payment.amount = float(data.get('amount') or payment.amount)
    if 'franchise_name' in data:
        payment.franchise_name = (data.get('franchise_name') or '').strip() or None
    if 'status' in data:
        payment.status = (data.get('status') or payment.status).strip().lower()
    if 'notes' in data:
        payment.notes = (data.get('notes') or '').strip() or None

    if payment.status != 'allocated' and payment.status != 'unmatched':
        payment.status = 'edited'

    db.session.commit()
    return jsonify(payment.to_dict()), 200


@api.get('/reports/summary')
@roles_required('admin')
def reports_summary():
    _seed_sample_payments()
    payments = PaymentTransaction.query.all()
    users = User.query.all()
    total_amount = sum(payment.amount for payment in payments)
    allocated_amount = sum(payment.amount for payment in payments if payment.status == 'allocated')
    unmatched_count = sum(1 for payment in payments if payment.status == 'unmatched')
    role_breakdown = {}
    for user in users:
        role = normalize_role(user.role)
        role_breakdown[role] = role_breakdown.get(role, 0) + 1

    return jsonify({
        'totals': {
            'users': len(users),
            'payments': len(payments),
            'total_amount': round(total_amount, 2),
            'allocated_amount': round(allocated_amount, 2),
            'unmatched_count': unmatched_count,
        },
        'role_breakdown': role_breakdown,
    }), 200

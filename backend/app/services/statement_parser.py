from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

import pdfplumber
from pypdf import PdfReader
from openpyxl import load_workbook

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None


DATE_PATTERN = re.compile(
    r"\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\b"
)
AMOUNT_PATTERN = re.compile(
    r"(?<!\d)(?:R\s*)?[+-]?(?:\d{1,3}(?:[ ,]\d{3})+|\d+)(?:[.,]\d{2})?(?:\s?(?:CR|DR))?(?!\d)",
    re.IGNORECASE,
)
REFERENCE_PATTERN = re.compile(
    r"\b(?:ref(?:erence)?[:\s-]*)?([A-Z0-9][A-Z0-9\-/.]{4,})\b",
    re.IGNORECASE,
)
SKIP_LINE_PATTERN = re.compile(
    r"opening balance|closing balance|balance brought forward|page \d+|statement period|account number|branch code|available balance|total charges|summary",
    re.IGNORECASE,
)
DEBIT_HINT_PATTERN = re.compile(
    r"\b(debit|dr\b|fee|charge|withdrawal|purchase|debit order|cash withdrawal|atm|transfer to|airtime|electricity|card purchase|payment to)\b",
    re.IGNORECASE,
)
CREDIT_HINT_PATTERN = re.compile(
    r"\b(credit|cr\b|deposit|payment|paid|salary|cash dep|transfer from|received|immediate payment|eft received)\b",
    re.IGNORECASE,
)
NAME_STOPWORDS = re.compile(
    r"\b(?:payment|deposit|transfer|received|credit|cash|immediate|eft|atm|ref|reference|from|by|via|trf|notification|online|banking)\b",
    re.IGNORECASE,
)

BANK_PATTERNS = {
    "nedbank": re.compile(r"\bnedbank\b", re.IGNORECASE),
    "absa": re.compile(r"\babsa\b", re.IGNORECASE),
    "fnb": re.compile(r"\bfirst national bank\b|\bfnb\b", re.IGNORECASE),
    "standard_bank": re.compile(r"\bstandard bank\b", re.IGNORECASE),
    "capitec": re.compile(r"\bcapitec\b", re.IGNORECASE),
}

HEADER_ALIASES = {
    "date": [
        "date", "transaction date", "posted date", "statement date", "process date", "value date", "txn date"
    ],
    "description": [
        "description", "transaction description", "details", "narrative", "memo", "transaction details", "entry description"
    ],
    "reference": [
        "reference", "payment reference", "our reference", "beneficiary reference", "customer reference", "statement reference", "ref no"
    ],
    "debit": ["debit", "withdrawal", "money out", "debits", "paid out"],
    "credit": ["credit", "deposit", "money in", "credits", "paid in"],
    "amount": ["amount", "transaction amount", "value", "amt"],
    "balance": ["balance", "running balance", "available balance"],
}

BANK_HEADER_HINTS = {
    "nedbank": {
        "date": ["processed date"],
        "description": ["transaction detail"],
        "reference": ["reference no", "other party reference"],
    },
    "absa": {
        "description": ["transaction description"],
        "reference": ["customer reference", "bank reference"],
    },
    "fnb": {
        "description": ["description", "transaction description"],
        "reference": ["ref", "reference"],
    },
    "standard_bank": {
        "description": ["narrative", "details"],
        "reference": ["document number", "reference"],
    },
    "capitec": {
        "description": ["description", "transaction type"],
        "reference": ["reference", "beneficiary reference"],
    },
}


@dataclass
class ParsedTransaction:
    payer_name: str
    reference: str
    amount: float
    franchise_name: str | None = None
    notes: str | None = None

    def to_dict(self):
        return {
            "payer_name": self.payer_name,
            "reference": self.reference,
            "amount": self.amount,
            "franchise_name": self.franchise_name,
            "notes": self.notes,
        }


class StatementParseError(ValueError):
    pass


class StatementParser:
    def parse_file(self, filename: str, file_bytes: bytes, franchise_name: str | None = None, bank_name: str | None = None) -> list[dict]:
        if not file_bytes:
            raise StatementParseError("The uploaded statement file is empty.")

        extension = Path(filename or "statement").suffix.lower()
        normalized_bank = self._normalize_bank_name(bank_name)

        if extension == ".pdf":
            rows = self.parse_pdf(file_bytes, franchise_name=franchise_name, bank_name=normalized_bank)
        elif extension == ".csv":
            rows = self.parse_csv(file_bytes, franchise_name=franchise_name, bank_name=normalized_bank)
        elif extension in {".xlsx", ".xlsm"}:
            rows = self.parse_xlsx(file_bytes, franchise_name=franchise_name, bank_name=normalized_bank)
        elif extension == ".xls":
            rows = self.parse_xls(file_bytes, franchise_name=franchise_name, bank_name=normalized_bank)
        elif extension == ".json":
            raise StatementParseError("JSON import should be sent as structured API data, not as an uploaded file.")
        else:
            raise StatementParseError("Unsupported statement file type. Upload PDF, CSV, XLSX, or XLS.")

        if not rows:
            raise StatementParseError(
                "No incoming payment transactions were found in this statement. Check the file layout or choose the correct bank."
            )
        return rows

    def parse_pdf(self, pdf_bytes: bytes, franchise_name: str | None = None, bank_name: str | None = None) -> list[dict]:
        lines = self._extract_candidate_lines(pdf_bytes)
        resolved_bank = bank_name or self._detect_bank_from_lines(lines)
        transactions = self._parse_pdf_lines(lines, franchise_name, resolved_bank)
        if not transactions:
            raise StatementParseError(
                "No payment transactions could be extracted from this PDF. Use a text-based bank statement PDF with visible transaction rows."
            )
        return [transaction.to_dict() for transaction in transactions]

    def parse_csv(self, file_bytes: bytes, franchise_name: str | None = None, bank_name: str | None = None) -> list[dict]:
        text = self._decode_text(file_bytes)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
        return [t.to_dict() for t in self._parse_tabular_rows(rows, franchise_name, bank_name)]

    def parse_xlsx(self, file_bytes: bytes, franchise_name: str | None = None, bank_name: str | None = None) -> list[dict]:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        rows: list[list[str]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append([self._stringify_cell(value) for value in row])
        return [t.to_dict() for t in self._parse_tabular_rows(rows, franchise_name, bank_name)]

    def parse_xls(self, file_bytes: bytes, franchise_name: str | None = None, bank_name: str | None = None) -> list[dict]:
        if xlrd is None:
            raise StatementParseError("XLS support requires the xlrd package to be installed.")
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        rows: list[list[str]] = []
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                rows.append([self._stringify_cell(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)])
        return [t.to_dict() for t in self._parse_tabular_rows(rows, franchise_name, bank_name)]

    def _parse_tabular_rows(self, rows: list[list[str]], franchise_name: str | None, bank_name: str | None) -> list[ParsedTransaction]:
        cleaned_rows = [[self._normalise_line(cell) for cell in row] for row in rows]
        header_index, header_map = self._detect_header_row(cleaned_rows, bank_name)
        if header_index is None or not header_map:
            raise StatementParseError(
                "Could not identify the statement columns. Expected columns like Date, Description, Reference, Credit, Debit, or Amount."
            )

        resolved_bank = bank_name or self._detect_bank_from_rows(cleaned_rows[: header_index + 5])
        transactions: list[ParsedTransaction] = []
        seen_keys: set[tuple[str, str, str]] = set()

        for row in cleaned_rows[header_index + 1 :]:
            if not any(row):
                continue
            tx = self._parse_tabular_record(row, header_map, franchise_name, resolved_bank)
            if not tx:
                continue
            key = (tx.reference.lower(), f"{tx.amount:.2f}", tx.payer_name.lower())
            if key in seen_keys:
                continue
            seen_keys.add(key)
            transactions.append(tx)

        return transactions

    def _detect_header_row(self, rows: list[list[str]], bank_name: str | None) -> tuple[int | None, dict[str, int]]:
        best_index = None
        best_map: dict[str, int] = {}
        best_score = 0

        alias_map = {k: list(v) for k, v in HEADER_ALIASES.items()}
        if bank_name and bank_name in BANK_HEADER_HINTS:
            for key, values in BANK_HEADER_HINTS[bank_name].items():
                alias_map.setdefault(key, []).extend(values)

        for idx, row in enumerate(rows[:20]):
            row_map = self._map_header_row(row, alias_map)
            score = len(row_map)
            if score > best_score and ("date" in row_map and ("description" in row_map or "reference" in row_map)):
                best_index = idx
                best_map = row_map
                best_score = score

        return best_index, best_map

    def _map_header_row(self, row: list[str], alias_map: dict[str, list[str]]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for idx, cell in enumerate(row):
            key = self._canonical_header(cell)
            if not key:
                continue
            for field, aliases in alias_map.items():
                if key == self._canonical_header(field) or key in {self._canonical_header(a) for a in aliases}:
                    mapped.setdefault(field, idx)
                    break
        return mapped

    def _canonical_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]", "", (value or "").strip().lower())

    def _parse_tabular_record(self, row: list[str], header_map: dict[str, int], franchise_name: str | None, bank_name: str | None) -> ParsedTransaction | None:
        date_text = self._row_value(row, header_map.get("date"))
        description = self._row_value(row, header_map.get("description"))
        reference = self._row_value(row, header_map.get("reference"))
        debit_text = self._row_value(row, header_map.get("debit"))
        credit_text = self._row_value(row, header_map.get("credit"))
        amount_text = self._row_value(row, header_map.get("amount"))

        if not date_text and not description and not reference:
            return None
        if date_text and not DATE_PATTERN.search(date_text) and not re.search(r"\d{2,4}", date_text):
            return None

        amount = self._resolve_tabular_amount(debit_text, credit_text, amount_text, description)
        if amount is None or amount <= 0:
            return None

        description = description or reference or ""
        if not description:
            return None
        if SKIP_LINE_PATTERN.search(description):
            return None

        resolved_reference = self._extract_reference(reference or description)
        if not resolved_reference:
            resolved_reference = self._make_reference_from_text(description, amount)

        payer_name = self._extract_payer_name(description, resolved_reference)
        if not payer_name:
            payer_name = resolved_reference or "Unknown payer"

        return ParsedTransaction(
            payer_name=payer_name[:255],
            reference=resolved_reference[:120],
            amount=float(amount),
            franchise_name=franchise_name,
            notes=f"Imported from {bank_name or 'bank'} tabular statement row",
        )

    def _resolve_tabular_amount(self, debit_text: str, credit_text: str, amount_text: str, description: str) -> Decimal | None:
        credit = self._parse_amount(credit_text) if credit_text else None
        debit = self._parse_amount(debit_text) if debit_text else None
        amount = self._parse_amount(amount_text) if amount_text else None

        if credit is not None and credit > 0:
            return credit
        if debit is not None and debit > 0 and credit is None:
            return None
        if amount is None:
            return None

        if amount > 0 and not self._looks_like_debit(description, amount_text):
            return amount
        if amount < 0:
            return None
        return None

    def _row_value(self, row: list[str], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return row[index].strip()

    def _detect_bank_from_rows(self, rows: list[list[str]]) -> str | None:
        joined = "\n".join(" ".join(row) for row in rows)
        return self._detect_bank_from_text(joined)

    def _detect_bank_from_lines(self, lines: list[str]) -> str | None:
        return self._detect_bank_from_text("\n".join(lines[:50]))

    def _detect_bank_from_text(self, text: str) -> str | None:
        for bank, pattern in BANK_PATTERNS.items():
            if pattern.search(text or ""):
                return bank
        return None

    def _normalize_bank_name(self, bank_name: str | None) -> str | None:
        if not bank_name:
            return None
        value = re.sub(r"[^a-z]", "_", bank_name.strip().lower())
        value = re.sub(r"_+", "_", value).strip("_")
        if value in {"standardbank", "standard_bank"}:
            return "standard_bank"
        if value in {"first_national_bank", "firstnationalbank"}:
            return "fnb"
        return value if value in BANK_PATTERNS else None

    def _extract_candidate_lines(self, pdf_bytes: bytes) -> list[str]:
        candidates: list[str] = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                candidates.extend(self._split_text_to_lines(text))
                for table in page.extract_tables() or []:
                    for row in table:
                        cells = [self._clean_cell(cell) for cell in row or [] if self._clean_cell(cell)]
                        if cells:
                            candidates.append(" ".join(cells))

        if not candidates:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                text = page.extract_text() or ""
                candidates.extend(self._split_text_to_lines(text))

        deduped: list[str] = []
        seen: set[str] = set()
        for raw_line in candidates:
            line = self._normalise_line(raw_line)
            if not line or line in seen:
                continue
            seen.add(line)
            deduped.append(line)
        return deduped

    def _split_text_to_lines(self, text: str) -> list[str]:
        return [line.strip() for line in text.splitlines() if line.strip()]

    def _clean_cell(self, value: object) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").strip()

    def _stringify_cell(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    def _normalise_line(self, line: str) -> str:
        line = line.replace("\xa0", " ")
        line = re.sub(r"\s+", " ", line).strip(" |\t")
        return line

    def _decode_text(self, file_bytes: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise StatementParseError("Could not decode the CSV file.")

    def _parse_pdf_lines(self, lines: Iterable[str], franchise_name: str | None, bank_name: str | None) -> list[ParsedTransaction]:
        results: list[ParsedTransaction] = []
        seen_keys: set[tuple[str, str, str]] = set()

        for line in lines:
            if not DATE_PATTERN.search(line):
                continue
            if SKIP_LINE_PATTERN.search(line):
                continue

            parsed = self._parse_pdf_line(line, franchise_name, bank_name)
            if not parsed:
                continue

            key = (parsed.reference.lower(), f"{parsed.amount:.2f}", parsed.payer_name.lower())
            if key in seen_keys:
                continue
            seen_keys.add(key)
            results.append(parsed)

        return results

    def _parse_pdf_line(self, line: str, franchise_name: str | None, bank_name: str | None) -> ParsedTransaction | None:
        amount_matches = list(AMOUNT_PATTERN.finditer(line))
        if not amount_matches:
            return None

        amount_match = self._pick_amount_match(amount_matches, line)
        amount = self._parse_amount(amount_match.group(0))
        if amount is None or amount <= 0:
            return None

        if self._looks_like_debit(line, amount_match.group(0)):
            return None

        description = self._extract_description(line, amount_match)
        if not description or len(description) < 3:
            return None

        reference = self._extract_reference(description)
        payer_name = self._extract_payer_name(description, reference)
        if not payer_name:
            payer_name = reference or "Unknown payer"
        if not reference:
            reference = self._make_reference_from_text(description, amount)

        return ParsedTransaction(
            payer_name=payer_name[:255],
            reference=reference[:120],
            amount=float(amount),
            franchise_name=franchise_name,
            notes=f"Imported from {bank_name or 'bank'} PDF statement row: {line[:300]}",
        )

    def _pick_amount_match(self, matches: list[re.Match], line: str) -> re.Match:
        if len(matches) == 1:
            return matches[0]

        line_lower = line.lower()
        has_balance_word = "balance" in line_lower
        if has_balance_word and len(matches) >= 2:
            return matches[-2]

        preferred = [m for m in matches if re.search(r"\b(?:cr|credit)\b", m.group(0), re.IGNORECASE)]
        if preferred:
            return preferred[-1]

        return matches[-1] if len(matches) == 2 else matches[-2]

    def _parse_amount(self, raw_amount: str) -> Decimal | None:
        if raw_amount is None:
            return None
        cleaned = str(raw_amount).upper().replace("R", "").replace(" ", "")
        if not cleaned:
            return None
        sign = Decimal("-1") if cleaned.endswith("DR") or cleaned.startswith("-") else Decimal("1")
        cleaned = cleaned.replace("CR", "").replace("DR", "")

        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif cleaned.count(",") == 1 and "." not in cleaned:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")

        try:
            value = Decimal(cleaned)
        except InvalidOperation:
            return None
        return value * sign

    def _looks_like_debit(self, line: str, amount_token: str) -> bool:
        if amount_token and (amount_token.strip().upper().endswith("DR") or amount_token.strip().startswith("-")):
            return True
        if DEBIT_HINT_PATTERN.search(line) and not CREDIT_HINT_PATTERN.search(line):
            return True
        return False

    def _extract_description(self, line: str, amount_match: re.Match) -> str:
        description = DATE_PATTERN.sub(" ", line)
        amount_token = re.escape(amount_match.group(0).strip())
        description = re.sub(rf"{amount_token}(?:\s+(?:R\s*)?[+-]?(?:\d{{1,3}}(?:[ ,]\d{{3}})+|\d+)(?:[.,]\d{{2}})?(?:\s?(?:CR|DR))?)?$", " ", description, flags=re.IGNORECASE)
        description = re.sub(r"\b(?:balance|credit|debit)\b.*$", " ", description, flags=re.IGNORECASE)
        description = re.sub(r"\s+", " ", description).strip(" -|")
        return description

    def _extract_reference(self, description: str) -> str:
        matches = [m.group(1) for m in REFERENCE_PATTERN.finditer(description or "")]
        if not matches:
            return ""
        matches.sort(key=len, reverse=True)
        return matches[0].upper()

    def _extract_payer_name(self, description: str, reference: str) -> str:
        payer = description or ""
        if reference:
            payer = re.sub(re.escape(reference), " ", payer, flags=re.IGNORECASE)
        payer = NAME_STOPWORDS.sub(" ", payer)
        payer = re.sub(r"[^A-Za-z0-9&./\- ]", " ", payer)
        payer = re.sub(r"\s+", " ", payer).strip(" -/")
        return payer.title()

    def _make_reference_from_text(self, description: str, amount: Decimal) -> str:
        compact = re.sub(r"[^A-Za-z0-9]", "", (description or "").upper())
        compact = compact[:8] or "PAYMENT"
        return f"{compact}-{int(amount * 100)}"
